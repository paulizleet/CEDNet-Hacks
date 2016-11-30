

from openpyxl import Workbook, load_workbook
import datetime
path = "C:\\Users\\pgallagherjr\\Documents\\"

def main():
	lines = get_lines()
	
	prd = get_products()
	
	print("Oooh lordy")
	print("This will take at worst " + str(len(lines) * len(prd)) + " iterations")
	print("here we goooooooooooo")
	
	listout = crunch_files(lines, prd)
	
	push_out(listout)
def get_products():
	"""whut"""
	
	f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.lsq", 'r')
	pl = []
	for line in f:
		toapp = []
		sp = line.split("|")
		toapp.append(sp[1].strip())
		toapp.append(sp[2].strip())
		toapp.append(0)
		toapp.append(0)
		pl.append(toapp)
	

	f.close()
	print(str(len(pl)) + " Products...")
	return pl	
	
def get_lines():
	f = open(path+"tblInvoiceDetailTable.txt")
	
	lots_of_lines = []
	for line in f:
		sp = line.split("|")
	#	for i,each in enumerate(sp):
	#		print(str(i) + ". " + each)
	#	input('...')
		toappend = []
		toappend.append(sp[17].strip())
		toappend.append(sp[18].strip())
		try:
			if sp[20][0] == "-":
				toappend.append(0)
			else:
				toappend.append(int(sp[20].strip()))
				
		except:
			toappend.append(0)
		#toappend.append(sp[20])
		
		
		lots_of_lines.append(toappend)
	
	print(str(len(lots_of_lines) ) + " Lines.")
	f.close()
	return lots_of_lines

def crunch_files(lines, prd):

	for line in lines:
#		print(line)
		for i,each in enumerate(prd):
			
			if line[:2] == each[:2]:
				prd[i][2] += 1
				prd[i][3] += line[2]
				break
	return prd

def push_out(list):
	wb = Workbook()
	ws = wb.active
	
	for each in list:
		ws.append(each)
	
	while True:
		try:
			wb.save("C:\\Users\\pgallagherjr\\Desktop\\crunched.xlsx")
			break
		except:
			input("close the spreadsheet idiot")
	
def binfreq():
	binwb = load_workbook("C:\\PaulScripts\\This Week's Stock Status.xlsx", data_only=True)
	binws = binwb.get_sheet_by_name("Sheet")

	print("got workbook")
	
	products = []
	
	for i in range(1, binws.max_row):
	
		if binws.cell(row=i, column=5).value is None:
			

			products.append([
				str(binws.cell(row=i, column=1).value).strip(),
				str(binws.cell(row=i, column=2).value).strip(),
				str(binws.cell(row=i, column=3).value).strip(),
				str(binws.cell(row=i, column=1).value).strip()]
				)						
			continue
			
		products.append([
			str(binws.cell(row=i, column=1).value).strip(),
			str(binws.cell(row=i, column=2).value).strip(),
			str(binws.cell(row=i, column=3).value).strip(),
			str(binws.cell(row=i, column=5).value).strip()]
			)			
			
	bins = []
	
	for each in products:
		if [each[3], 0] in bins:
			continue
		else:
			bins.append([each[3], 0])
	nextbins = bins
	wb = load_workbook("C:\\PaulScripts\\3moavg.xlsx", data_only=True)
	ws = wb.active
	
	
	for i in range(2, ws.max_row):
		mfrprd = [ws.cell(row=i, column=1).value.strip(), 
					str(ws.cell(row=i, column = 2).value).strip()]
		#input(mfrprd)
		for each in products:
			if each[:2] == mfrprd:
				#input(each)
				#input(ws.cell(row=i,column=8).value)
				avg = ws.cell(row=i, column = 8).value
				for i in range(0, len(bins)):
					if bins[i][0] == each[3]:
						bins[i][1] += int(avg)
						
	
	ws = wb.create_sheet("bin avgs")
	
	for each in bins:
		ws.append(each)
		
	wb.save("C:\\PaulScripts\\3moavg.xlsx")
	
	print("saved 3moavg to PaulScripts")
	print("Continuing to saleyear.xlsx")
	
	
	#wb.close()
	
	#wb = load_workbook("C:\\Users\\pgallagherjr\\Desktop\\saleyear.xlsx", read_only = True)
	#ws = wb.active
	
	#input(ws.max_row)
	f = open("C:\\Users\\pgallagherjr\\Desktop\\saleyear.txt", encoding="utf8")
	stripped = []
	
	for i, line in enumerate(f):
		s = line.split("\t")
		

	
		if i % 1000 != 0:
			#print(i)
			
			stripped.append([s[0], s[1]])
			
			continue
		print(i)
		print("pulled mfr/prd")
		
		for mfrprd in stripped:
			#print(i)

			for each in products:
				if each[:2] == mfrprd:
					for i in range(0, len(nextbins)):
							if bins[i][0] == each[3]:
								bins[i][1] += 1
		stripped = []
		#print(stripped)
		
		
		
	outbook = Workbook()
	outsheet = outbook.active
	
	for each in nextbins:
		outsheet.append(each)
		
		
	outbook.save("C:\\PaulScripts\\yeartotals.xlsx")
	
	

	
if __name__ == "__main__":
	now = datetime.datetime.now()

	binfreq()
	
	print(datetime.datetime.now() - now)