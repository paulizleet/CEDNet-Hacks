from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime

def main():

	#spksdata = speaks()
	
	#prod = cednet_products()
	
	mash_workbook(cednet_products(), speaks())
	
def mash_workbook(cednet, spks):
	#print(spks.__len__())
	#print(spks[0])
	#print(cednet[0])
	wb = Workbook()
	ws = wb.create_sheet()
	
	ws.cell(row=1, column = 1).value = ("Mfr")
	ws.cell(row=1, column = 2).value = ("Cat #")
	ws.cell(row=1, column = 3).value = ("Desc")
	ws.cell(row=1, column = 4).value = ("OH Quantity")
	ws.cell(row=1, column = 5).value = ("Carton Quantity")
	ws.cell(row=1, column = 6).value = ("Bin Location")
	ws.cell(row=1, column = 7).value = ("# Sold")
	ws.cell(row=1, column = 8).value = ("# Orders")

	ced = []
	for i in range(0, cednet.__len__()-1):
		line =	cednet[i]
		#line.append(cednet[i])
		for each2 in spks:
			if each2[1:3] == cednet[i][:2]:
				line.append(each2[7])
				line.append(each2[8])
				break
				
		ced.append(line)



	
	for i in range(0, ced.__len__()):
		for j in range(0, ced[i].__len__()):
			print(ced[i][j])
			ws.cell(row=i+2, column=j+1).value = ced[i][j]
	
	wb.save("C:\Documents and Settings\pgallagherjr\Desktop\speak.xlsx")

	
	
def cednet_products():
	f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.lsq")
	products = []
	
	prod = []
	for line in f:
		prod = []
		l = line.split("|")
		prod.append(l[1].strip())
		prod.append(l[2].strip())
		prod.append(l[5].strip())
		prod.append(l[15].strip())
		prod.append(l[29].strip())
		
		products.append(prod)
	
	return products
	

def speaks():

	f = open("C:\\Invsys\\Algorithm\\SPKMFRDT.lsq")

	mfrs = []


	for line in f:
		mfrs.append(line[3:8].strip())
		#print(line[3:8])

	f = open("C:\\Invsys\\Algorithm\\SPKCUSDT.lsq")
	customers = []

	for line in f:
		l = line.split("|")
		customers.append(l[4])
		
	customers.append(None)
		
	f = open("C:\\Users\pgallagherjr\Desktop\product.txt")

	lines = []

	for line in f:
		#input("")
		
		#print( line.split("\t"))
		#print( line.split(" "))

		if line[:line.find("\t")].strip() in mfrs:
			#print("OK")
			lines.append(line)
			continue
		if line[:line.find("\t")].strip() in customers:
			#print("OK")
			lines.append(line)


	#print(customers)
	cs = ""

	lines2 = []
	for i, line in enumerate(lines):
		
		#	print(line)
		#	print(line[:line.find("\t")].strip())
		#	input("...")
			
		if line[:line.find("\t")].strip() in customers:
			cs	= line[:line.find("\t")].strip()
			#print(cs)
			lines.pop(i)
			continue

		lines2.append(cs + "\t" + line)
		
	final = []

	for i, each in enumerate(lines2):
		final.append(each.split("\t"))
		for j, each2 in enumerate(final[i]):
			final[i][j] = each2.strip()


	print(final[0])
			

			
	duration = datetime.now()

	forlater = final
			
	#Customer and product
	combined = []
	mistakes = 0
	items = 0
	for each in final:
		
		items += 1
		if items%1000 == 0:
			print(str(items / final.__len__()))
		try:
			#try:
				#print(each)
			#except UnicodeEncodeError:
			#	pass
				
			asdf = False
			if combined.__len__() > 0:
			
				
					
				'''for entry in combined:
					if each[:3] == entry[:3]:
					
						entry[7] = int(entry[7]) + int(each[7])
						asdf = True
						break'''
				#print("--------------------------------------------------")
				#input('...')
				#print("combined len: " + str(combined.__len__()))
				
				for i in range(combined.__len__(), 1, -1):
					#print(">:-)")
					#print(combined[i-1])
					#print("8-)")
					if each[0] != combined[i-1][0]:
					#	print("customer done")
						break
					
					if each[:3] == combined[i-1][:3]:
					
						combined[i-1][8]+=1
						combined[i-1][7] = int(combined[i-1][7]) + int(each[7])
						asdf = True
						break
						
			#print("x)")			
			
			
			try:
				each[7] = int(each[7])
			except ValueError:
			#	print("adding a value")
				each.insert(7, 0)

			#print("=o")
			
			if asdf == False:
				each.append(1)
				combined.append(each)
		except IndexError:
			#print("mistake...")
			mistakes +=1
			pass
			

	f = open("C:\\Users\pgallagherjr\Desktop\Customer_Product.txt", "w")

	for i, each in enumerate(combined):
		ln = str(i)
		for each2 in each:
			ln = ln + "\t" + str(each2)
		ln = ln + "\n"
		f.write(ln)
		
	f.close()


	print(str(items) + " items.")
	print(str(mistakes) + " mistakes.")
			
	print(datetime.now() - duration)
	duration = datetime.now()





	#only products
	final = []
	mistakes = 0
	items = 0
	for each in forlater:
		items += 1
		
		if items%1000 == 0:
			print(str(items / forlater.__len__()))
		try:
			#try:
				#print(each)
			#except UnicodeEncodeError:
			#	pass

			asdf = False
			if final.__len__() > 0:
					
				for entry in final:
					if each[2] == entry[2]:
					
						entry[7] = int(entry[7]) + int(each[7])
						entry[8] += 1
						asdf = True
						break
			try:
				each[7] = int(each[7])
			except ValueError:
				each.insert(7, 0)

			
			if asdf == False:
				each.append(0)
				final.append(each)

				
		except IndexError:
			mistakes +=1
			pass
			

			
	'''f = open("C:\Documents and Settings\pgallagherjr\Desktop\Product.txt", "w")

	for i, each in enumerate(final):
		ln = str(i)
		for each2 in each:
			ln = ln + "\t" + str(each2)
		ln = ln + "\n"
		f.write(ln)
		
	f.close()'''
	
	print(str(items) + " items.")
	print(str(mistakes) + " mistakes.")
	print(datetime.now() - duration)
	return final


main()