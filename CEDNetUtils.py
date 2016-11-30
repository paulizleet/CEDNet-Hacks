#CEDNetUtils.py
#Usually I copy/paste code from one script to the next.
#I'll put all common code in here instead.

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import dimensions
from openpyxl.styles import Side, Border
from datetime import datetime
import math
import random

def split_product_lines():

	try:
		f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.lsq")
		
	except FileNotFoundError:
		print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
		print("It can be found at Maintainance > Product > Data Files")
		return
	
	prd = []
	splits = []
	for line in f:
		splits.append(line.split("|"))
	return splits

def get_mfrs():
	mfr = []
	
	try:
		f = open("C:\\Invsys\\Algorithm\\SPKMFRDT.lsq")
		
	except FileNotFoundError:
		print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
		print("It can be found at Maintainance > Product > Data Files")
		return
		
	for line in f:
		mfr.append(line.split("|"))
	return mfr
	
	
def get_products():

	prd = []

	try:
		f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.lsq")
		
	except FileNotFoundError:
		print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
		print("It can be found at Maintainance > Product > Data Files")
		return
	
	prd = []
	splits = []
	for line in f:
		splits.append(line.split("|"))
	for line in splits:
		p = []
		
		p.append(line[1].strip()) #mfr
		p.append(line[2].strip()) #cat num
		p.append(line[5].strip()) #Description
		p.append(line[15].strip()) #onhand qty
		p.append(line[29].strip()) #bin location
		prd.append(p)
		
	return prd
	
def get_customers():
	try:

		cs = open("C:\\Invsys\\Algorithm\\SPKCUSDT.lsq")

		customers = []
		cline = []
		for line in cs:
			cline = []
			#print(line)
			l = line.split("|")
			cline.append(l[2]) #ACCT
			print("|" + l[2] + "|")
			cline.append(l[3]) #NAME
			cline.append((l[5]+ " " + l[6]).strip()) #Addr
			cline.append(l[9]) #city
			cline.append(l[7]) #zip
			cline.append(l[8]) #State


			customers.append(cline)

		correct = False

		print("0. Account Num: "		+	customers[34][0])
		print("1. Customer Name: "	+	customers[34][1])
		print("2. Customer Addr:"		+	customers[34][2])
		print("3. Customer City:"		+	customers[34][5])
		print("4. Customer State: "	+	customers[34][4])
		print("5. Customer Zip: "		+	customers[34][3])

		correct = input("Is this correct?  y/n")

		if correct.lower() == "n":
			print("Please fix the values named CUST_X")
			return False


	except FileNotFoundError:
		print("CEDNet data files not found.  \nPlease run the Data Files export function in CEDNet")
		print("It can be found at Maintainance > Product > Data Files")


	return customers

	
def write_excel_sheet(header = None, list = None, sheet=None, border=None):
	
	offset = 1
	if header != None:
		sheet.append(header)
		offset = 2
		
	for i, each in enumerate(list):
		sheet.append(each)
		if border != None:
			for j in range(1, sheet.max_column):
				sheet.cell(row=i+offset, column=j).border = border
	return sheet


		
	

	