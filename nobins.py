#NOBINS.PY

'''
	This script reads the CEDNet data files and turns out an excel spreadsheet with all products
	that don't have a bin location in the system.  It also gives you a sheet with all
	of the products in order
	for you to help punch them in using ShelfAddresses.py

'''
from openpyxl import Workbook
#from openpyxl import load_workbook
from openpyxl.styles import Side, Border
import CEDNetUtils as CED
#from openpyxl.formatting import Rule
#from openpyxl.worksheet import dimensions

CELL_FORMULA = "=IF(AND(G{s}<>\"\", H{s}<>\"\"), UPPER(CONCATENATE(G{s}, \".\"" \
					",H{s}, IF(I{s}<>\"\", CONCATENATE(\".\", I{s}),\"\"))),\"\")"
def run():

	""" Run the script.	 Needs a method and the name block depending on how it gets ran"""

	while True:
		print("This will overwrite the current stock status workbook\nDo you wish to continue? Y/n")
		choice = input("??")
		if choice.lower() == "y":
			break
		if choice.lower() == "n":
			return
		print("Invalid choice.")

	product_list = CED.get_products()




	workbook = Workbook()
	worksheet = workbook.get_sheet_by_name("Sheet")
	worksheet.column_dimensions['A'].width = 8
	worksheet.column_dimensions['B'].width = 24
	worksheet.column_dimensions['C'].width = 40
	worksheet.column_dimensions['D'].width = 10
	worksheet.column_dimensions['E'].width = 10
	side = Side(border_style='thin', color="00000000")
	border = Border(left=side,right=side,bottom=side, top=side)


	header = ["MFR", "CAT #", "DESCRIPTION", "OH Qty.", "Current Bin", "New Bin"]
	worksheet = CED.write_excel_sheet(header = header, list = product_list, sheet = worksheet, border = border)

	nobinsheet = workbook.create_sheet('nobins', 1)
  #	 dim = dimensions.ColumnDimension(nbs, bestFit=True)

	dontcount = ['CR', 'PECO', 'ZZ588', 'PV6-2KV', 'C588', 'WIRE', 'COND',
				 'CORD', 'DYSON', 'FLEX', 'LIQ', 'HYU', 'LG', 'FRO', 'POWON',
				 'OMNI', 'SWLD', "310168C", "310168D", "310760", "320123M", "320168M",
				 "51-7556-056H", "XR-10-132A", "XR-10-168A", "XR-10-204A",
				 "XR-100-132B", "XR-100-168A", "XR-100-168B", "XR-100-204B",
				 "XR-1000-168B", "G134OS1", "G582OS1", "PHONO", "JINKO-JKM260P-60",
				 "NLHBLEV20013", "QCELL", "C3410", "A1200HS10PG", "TESLA", "004300C", "CANS"]

	nobins = []
	for j, each in enumerate(product_list):

		if each[0].strip() in dontcount or each[1].strip() in dontcount:
		
			worksheet.cell(row=j+2, column=5).value = "n/a"
			continue

		if each[-1] == '' or each[-1] is None:

			nobins.append(each)


	header = ["MFR", "CAT #", "DESCRIPTION", "OH Qty.", "New Bin"]

	nobinsheet=CED.write_excel_sheet(header = header, list = nobins, sheet=nobinsheet, border = border)
	nobinsheet.column_dimensions['A'].width = 8
	nobinsheet.column_dimensions['B'].width = 24
	nobinsheet.column_dimensions['C'].width = 40
	nobinsheet.column_dimensions['D'].width = 10
	nobinsheet.column_dimensions['E'].width = 10
	workbook.save("C:\\PaulScripts\\This Week's Stock Status.xlsx")


	print("Spreadsheet saved to C:\\PaulScripts\\This Week's Stock Status.xlsx")
	
	print("\nThere are " + str(len(nobins)) + " items without bin locations.")

if __name__ == "__main__":
	run()
else:
	print("imported nobins.py")
