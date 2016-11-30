#panelstock

from openpyxl import Workbook
from openpyxl.worksheet import dimensions

import datetime
import CEDNetUtils as CED

mfrs = ["LG", "HYU", "SWLD"]
def main():
	
	lns = get_lines()
	print(lns)
	write_book(lns)
	
	
def get_lines():

	f = open("C:\\Invsys\\Algorithm\\SPKPRDDT.LSQ")

	pgs = []

	for line in f:
		ln = line.split("|")
		if ln[1].strip() in mfrs:
			pgs.append(ln)
	
	return pgs
					
def write_book(lns):
	wb = Workbook()
	ws = wb.active
	ws.column_dimensions['A'].width = 7
	ws.column_dimensions['B'].width = 23
	ws.column_dimensions['C'].width = 40
	ws.column_dimensions['D'].width = 10
	ws.column_dimensions['E'].width = 10
	
	ws.append(["MFR", "CAT #", "DESC", "OH QTY", "AVAIL."])
	
	for each in lns:
		ws.append([
			each[1].strip(), 
			each[2].strip(),
			each[5].strip(), 
			each[15].strip(),
			each[31].strip()
			])		
	ws.cell("F1").value = "Updated:"
	ws.cell("G1").value = datetime.date.today()
	wb.save("C:\\Users\\pgallagherjr\\Dropbox\\Panel Stock\panels.xlsx")
	
	
	
main()
