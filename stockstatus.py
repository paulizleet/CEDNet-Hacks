from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime



def main():
	f = open("C:\\Documents and Settings\\pgallagherjr\\Desktop\\document.txt")
	toskip = [1, 2, 3, 4, 5, 6, 63]
	
	skip = 0
	lines = []
	for line in f:
		skip+=1
		if skip in toskip:
			if skip == 63:
				skip =0
				
			continue
		lines.append(line.split("  "))
		
	print(lines[0])
	print(lines[0][:5])
	print(lines[0][5:])
	
	firstlines = []
	secondlines = []
	
	for each in lines:
		firstlines.append(each[:5])
		secondlines.append(each[5:])
	
	
	lines = firstlines + secondlines
	wb = Workbook()
	ws = wb.active
	
	for i,dach in enumerate(lines):
		each = dach
		while True:
			if each.__len__() < 5:
				each.append(0)
				continue
			break
			
		ws.cell(row=i+1, column = 1).value = each[0]
		ws.cell(row=i+1, column = 2).value = each[1]
		ws.cell(row=i+1, column = 3).value = each[2]
		ws.cell(row=i+1, column = 4).value = each[3]
		ws.cell(row=i+1, column = 5).value = each[4]
		
	wb.save("C:\\Documents and Settings\\pgallagherjr\\Desktop\\stockstatusreport.xlsx")
			
	

main()